"""Utilities for handling input test files"""
import re

from openpyxl import load_workbook

from city import City
from city2 import City2


class TestData:
    """Class representing testcase data parsed from input files"""
    def __init__(self, truck_count: int, capacity: int, cities: list[City], optimal: float, solution: list[list[int]]) -> None:
        self.truck_count = truck_count
        self.capacity = capacity
        self.cities = cities
        self.optimal = optimal
        self.solution = solution

    def __str__(self) -> str:
        cities = '\n\t'.join([F'{i}. {city}' for i, city in enumerate(self.cities)])
        solution = '\n\t'.join([str(path) for path in self.solution])
        return F'''\
Truck count: {self.truck_count}
Capacity: {self.capacity}
Cities: {cities}
Optimal value: {self.optimal}
Optimal:{solution}'''

class CVRPTestParser:
    """Class parsing testcase files to TestData"""
    @classmethod
    def parse(cls, test_name: str) -> TestData:
        """Static method for parsing test files into instance of TestData class"""
        test_path = '../testsets/' + test_name
        with open(test_path + '.vrp', 'r', encoding='utf-8') as test_file:
            test_content = test_file.read()
            truck_count = re.search(r"No of trucks: (\d+)", test_content, re.MULTILINE).group(1)
            capacity = re.search(r"^\s?CAPACITY\s?: (\d+)\s?$", test_content, re.MULTILINE).group(1)
            positions = re.findall(r"^\s?(\d+) (\d+) (\d+)\s?$", test_content, re.MULTILINE)
            demand = re.findall(r"^\s?(\d+) (\d+)\s?$", test_content, re.MULTILINE)
            cities = []
            for position in positions:
                for dem in demand:
                    if position[0] == dem[0]:
                        cities.append(City(int(position[1]), int(position[2]), int(dem[1])))

        with open(test_path + '.sol', 'r', encoding='utf-8') as solution_file:
            solution_content = solution_file.read()
            optimal = re.search(r"Cost (\d+)", solution_content, re.MULTILINE).group(1)
            raw_sol = re.findall(r"^\s?Route #(\d+)\s?:\s?(.*)\s?$", solution_content, re.MULTILINE)
            paths = [[int(vertex) for vertex in unparsed[1].split()] for unparsed in raw_sol]

        test_data = TestData(int(truck_count), int(capacity), cities, int(optimal), paths)
        return test_data


class CVRPTestParser2:
    """Class parsing city data from an xlsx file to TestData"""
    @classmethod
    def parse(cls, file_path: str) -> TestData:
        """Static method for parsing city data from xlsx file into instance of TestData class"""
        cities = []
        truck_count = 32
        capacity = 8
        optimal = 0
        paths = []

        try:
            workbook = load_workbook(file_path, read_only=True)
            sheet = workbook.active

            for row in sheet.iter_rows(min_row=2, values_only=True):
                name, longitude, latitude, demand = row[0], row[1], row[2], row[3]
                cities.append(City2(name, longitude, latitude, demand))

        except Exception as e:
            print(f"Error parsing xlsx file: {e}")

        test_data = TestData(truck_count, capacity, cities, optimal, paths)
        return test_data


class CVRPTestParser3:
    """Class parsing city and truck data from system input to TestData"""

    @classmethod
    def parse(cls, garbage_collection_points: list, garbage_collection_points_longitude: list,
              garbage_collection_points_latitude: list, garbage_amounts: list,
              truck_capacity: int, transfer_station_name: str, transfer_station_longitude: float,
              transfer_station_latitude: float) -> 'TestData':
        """Static method for parsing system input data into instance of TestData class"""
        cities = []
        # Insert transfer station as the first city
        transfer_station = City2(transfer_station_name, transfer_station_longitude, transfer_station_latitude, 0)
        cities.append(transfer_station)

        # Add garbage collection points as cities
        for i in range(len(garbage_collection_points)):
            city = City2(garbage_collection_points[i], garbage_collection_points_longitude[i],
                         garbage_collection_points_latitude[i], garbage_amounts[i])
            cities.append(city)

        # Calculate truck count based on truck capacity and total garbage amount
        total_garbage = sum(garbage_amounts)
        truck_count = (total_garbage + truck_capacity - 1) // truck_capacity

        return TestData(truck_count, truck_capacity, cities, optimal=0, solution=[])
